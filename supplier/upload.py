# supplier/upload.py
import logging
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from .models import Supplier, Technology, Logistic
from .forms import UploadExcelForm

# Инициализация логгера для текущего модуля
logger = logging.getLogger(__name__)

# Конфигурация для каждой модели
UPLOAD_CONFIG = {
    'supplier': {
        'model': Supplier,
        'template': 'upload/upload_suppliers.html',
        'redirect_url': 'upload_suppliers',
        'log_prefix': 'Supplier'
    },
    'technology': {
        'model': Technology,
        'template': 'upload/upload_technology.html',
        'redirect_url': 'upload_technology',
        'log_prefix': 'Technology'
    },
    'logistic': {
        'model': Logistic,
        'template': 'upload/upload_logistic.html', # Исправлено
        'redirect_url': 'upload_logistic',
        'log_prefix': 'Logistic'
    }
}

# Индексы обязательных полей (после id)
REQUIRED_FIELD_INDICES = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11] # index, country, ..., email

# Сопоставление полей модели с индексами в Excel (после id)
FIELD_MAPPING = {
    'index': 1,
    'country': 2,
    'category': 3,
    'name': 4,
    'website': 5,
    'description': 6,
    'product': 7,
    'contact': 8,
    'description_ru': 9,
    'product_ru': 10,
    'email': 11,
    'tn_ved': 12,
    'price': 13,
    'price_date': 14,
    'created_date': 15,
    'updated_date': 16,
}

def parse_date(date_str):
    """Парсит строку даты в объект date."""
    if date_str is None:
        return None
    if isinstance(date_str, datetime):
        return date_str.date()
    if isinstance(date_str, str):
        try:
            return datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            logger.warning(f"Неверный формат даты: {date_str}")
            return None
    logger.warning(f"Неожиданный тип даты: {type(date_str)} - {date_str}")
    return None

def parse_float(value, default=10.0):
    """Парсит значение в float."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        logger.warning(f"Неверный формат float: {value}")
        return default

def parse_int(value):
    """Парсит значение в int."""
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        logger.warning(f"Неверный формат int: {value}")
        return None


def process_excel_file(ws, model_class, config_key, request): # Добавлен request как аргумент
    """Обрабатывает данные из листа Excel и возвращает список объектов для создания и список ошибок."""
    headers = [cell.value for cell in ws[1]]
    required_headers = [
        "id", "index", "country", "category", "name", "website", "description",
        "product", "contact", "description_ru", "product_ru", "email", "tn_ved",
        "price", "price_date", "created_date", "updated_date",
        # "search_vector_product", "search_vector_product_ru", # Эти поля обычно заполняются в БД
    ]

    # Проверяем наличие всех обязательных заголовков (независимо от порядка)
    if not set(required_headers).issubset(set(headers)):
        missing = set(required_headers) - set(headers)
        raise ValueError(f"Отсутствуют обязательные заголовки: {missing}")

    items_to_create = []
    errors = [] # Список для накопления сообщений об ошибках строк
    errors_count = 0
    processed_count = 0

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        processed_count += 1
        row_data = list(row) # Преобразуем в список для удобства

        # Проверка обязательных полей
        skip_row = False
        for idx in REQUIRED_FIELD_INDICES:
            if not row_data[idx]: # Проверяем на None или пустую строку
                error_msg = f"Ошибка в строке {row_num}: пустое поле '{headers[idx]}'"
                logger.error(f"{config_key.capitalize()} - {error_msg}")
                errors.append(error_msg) # Добавляем ошибку в список
                errors_count += 1
                skip_row = True
                break # Прекращаем проверку этой строки
        if skip_row:
            continue

        try:
            # Создаем словарь данных для объекта модели
            data_dict = {}
            for field_name, excel_index in FIELD_MAPPING.items():
                raw_value = row_data[excel_index]
                if field_name in ['index']:
                    data_dict[field_name] = parse_int(raw_value)
                elif field_name in ['price']:
                    data_dict[field_name] = parse_float(raw_value)
                elif field_name in ['price_date', 'created_date', 'updated_date']:
                    data_dict[field_name] = parse_date(raw_value)
                else: # Остальные поля - строки
                    data_dict[field_name] = str(raw_value) if raw_value is not None else ""

            # Создаем экземпляр модели
            item = model_class(**data_dict)
            items_to_create.append(item)

        except Exception as e:
            error_msg = f"Ошибка обработки строки {row_num}: {e}"
            logger.error(f"{config_key.capitalize()} - {error_msg}")
            errors.append(error_msg) # Добавляем ошибку в список
            errors_count += 1

    return items_to_create, processed_count, errors_count, errors # Возвращаем также список ошибок


def _generic_upload_view(request, config_key):
    """Универсальная функция загрузки."""
    config = UPLOAD_CONFIG.get(config_key)
    if not config:
        messages.error(request, "Неверная конфигурация загрузки.")
        return redirect('main') # Или другая страница по умолчанию

    model_class = config['model']
    template_name = config['template']
    redirect_url = config['redirect_url']
    log_prefix = config['log_prefix']

    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                excel_file = request.FILES["excel_file"]
                logger.info(f"{log_prefix} - Начало загрузки файла: {excel_file.name}")

                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active

                # Передаем request в process_excel_file
                items_to_create, processed_count, errors_count, errors = process_excel_file(ws, model_class, config_key, request)

                # Отправляем сообщения об ошибках строк пользователю
                # Ограничим количество сообщений, чтобы не перегружать UI
                max_error_messages = 20 # Можно настроить
                for error_msg in errors[:max_error_messages]:
                    messages.error(request, error_msg)
                if len(errors) > max_error_messages:
                    messages.error(request, f"... и ещё {len(errors) - max_error_messages} ошибок.")

                success_count = 0
                if items_to_create:
                    # Пакетная загрузка
                    batch_size = 1000
                    total_batches = (len(items_to_create) + batch_size - 1) // batch_size
                    for i in range(0, len(items_to_create), batch_size):
                        batch = items_to_create[i:i + batch_size]
                        try:
                            with transaction.atomic():
                                created_objs = model_class.objects.bulk_create(batch, ignore_conflicts=False)
                                success_count += len(created_objs)
                                logger.info(f"{log_prefix} - Загружено пакет {i//batch_size + 1}/{total_batches}, {len(created_objs)} записей.")
                        except Exception as e:
                            logger.error(f"{log_prefix} - Ошибка при загрузке пакета {i//batch_size + 1}: {e}")
                            messages.error(request, f"Ошибка при загрузке части данных (пакет {i//batch_size + 1}): {e}")
                            # Можно прервать загрузку или продолжить, зависит от требований
                            # break

                messages.success(
                    request,
                    f"Обработано строк: {processed_count}. "
                    f"Успешно загружено: {success_count}. "
                    f"Ошибок: {errors_count}."
                )
                logger.info(f"{log_prefix} - Завершена загрузка. Обработано: {processed_count}, Загружено: {success_count}, Ошибок: {errors_count}")
                return redirect(redirect_url)

            except ValueError as ve: # Ошибки валидации заголовков
                logger.error(f"{log_prefix} - Ошибка валидации файла: {ve}")
                messages.error(request, f"Ошибка валидации файла: {ve}")
            except Exception as e:
                logger.error(f"{log_prefix} - Критическая ошибка загрузки файла: {e}", exc_info=True)
                messages.error(request, f"Критическая ошибка при обработке файла: {e}")

            return render(request, template_name, {"form": form})

    else:
        form = UploadExcelForm()
    return render(request, template_name, {"form": form})


def upload_suppliers(request):
    """Обрабатывает загрузку данных о поставщиках."""
    return _generic_upload_view(request, 'supplier')

def upload_technology(request):
    """Обрабатывает загрузку данных о технологиях."""
    return _generic_upload_view(request, 'technology')

def upload_logistic(request):
    """Обрабатывает загрузку данных о логистике."""
    return _generic_upload_view(request, 'logistic')

# TODO: вспомогательные функции для возможного использования 
def export_to_excel(request):
    '''Скачать базу поставщиков в xlsx'''
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') # Исправлено
    response['Content-Disposition'] = 'attachment; filename="suppliers.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    # Добавить заголовки
    headers = [
        "id", "index", "country", "category", "name", "website", "description",
        "product", "contact", "description_ru", "product_ru", "email", "tn_ved",
        "price", "price_date", "created_date", "updated_date"
    ]
    ws.append(headers)

    # Добавить данные из модели
    items = Supplier.objects.all().values(*headers) # Используем values для эффективности
    for item_data in items:
        # Обработка дат, если они не сериализуются автоматически
        row = []
        for h in headers:
            val = item_data[h]
            if isinstance(val, datetime):
                val = val.date() # Или val.strftime('%Y-%m-%d')
            row.append(val)
        ws.append(row)

    # Сохранить рабочую книгу в HttpResponse
    wb.save(response)
    return response

# --- Функции удаления ---
# Рассмотрите возможность добавления подтверждения удаления (например, через POST-запрос)
def supplier_delete(request):
    if request.method == "POST": # Рекомендуется использовать POST для удаления
        count, _ = Supplier.objects.all().delete()
        messages.success(request, f"Удалено {count} записей из поставщиков.")
        return redirect('main') # Или другая страница после удаления
    else:
        # Отобразить страницу подтверждения
        return render(request, 'upload/confirm_delete.html', {'model_name': 'Поставщики'})

def technology_delete(request):
    if request.method == "POST":
        count, _ = Technology.objects.all().delete()
        messages.success(request, f"Удалено {count} записей из технологий.")
        return redirect('main')
    else:
        return render(request, 'upload/confirm_delete.html', {'model_name': 'Технологии'})

def logistic_delete(request):
    if request.method == "POST":
        count, _ = Logistic.objects.all().delete()
        messages.success(request, f"Удалено {count} записей из логистики.")
        return redirect('main')
    else:
        return render(request, 'upload/confirm_delete.html', {'model_name': 'Логистика'})
